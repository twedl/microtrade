"""Tests for the Excel -> Spec converter and the `import-spec` CLI."""

from __future__ import annotations

from pathlib import Path

import pytest
from typer.testing import CliRunner

from microtrade import schema
from microtrade.cli import app
from microtrade.config import WorkbookConfig, load_config
from microtrade.excel_spec import derive_workbook_id, normalize_dtype, read_workbook
from microtrade.schema import TRADE_TYPES, SpecError, load_spec
from tests._helpers import (
    SHEET_TITLES,
    build_project_config,
    build_workbook,
    default_filename_pattern,
)


def test_normalize_dtype_handles_common_aliases() -> None:
    assert normalize_dtype("string") == "Utf8"
    assert normalize_dtype(" INT ") == "Int64"
    assert normalize_dtype("float64") == "Float64"
    assert normalize_dtype("date") == "Date"
    assert normalize_dtype("Char") == "Utf8"
    assert normalize_dtype("Num") == "Int64"


def test_normalize_dtype_rejects_unknown() -> None:
    with pytest.raises(SpecError, match="unrecognized dtype"):
        normalize_dtype("chronology")


def test_read_workbook_produces_spec_per_trade_type(
    schema_workbook: Path, workbook_config: WorkbookConfig
) -> None:
    specs = read_workbook(schema_workbook, workbook_config)
    assert set(specs) == set(TRADE_TYPES)

    imports = specs["imports"]
    assert imports.trade_type == "imports"
    assert imports.effective_from == "2020-01"
    assert imports.record_length == 53
    assert [c.physical_name for c in imports.columns] == [
        "period",
        "hs_code",
        "country_coo",
        "district_entry",
        "value_usd",
        "qty_kg",
    ]
    assert imports.source is not None
    assert imports.source.workbook == schema_workbook.name
    assert imports.source.sheet == "ImportsSheet"
    assert imports.source.filename_pattern == default_filename_pattern("ImportsSheet")
    assert imports.derived == (("year", "year(period)"), ("month", "month(period)"))

    exports_nonus = specs["exports_nonus"]
    assert [c.dtype for c in exports_nonus.columns] == ["Date", "Utf8", "Utf8", "Float64"]


def test_read_workbook_rejects_missing_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("only_one")
    ws.append(["Position", "Description", "Length", "Type"])
    ws.append([1, "a", 5, "Char"])
    wb_path = tmp_path / "wb.xlsx"
    wb.save(wb_path)

    # Config references a sheet that isn't in the workbook.
    config_path = build_project_config(tmp_path / "microtrade.yaml", wb_path, "2024-01")
    workbook_config = load_config(config_path).get_workbook(wb_path)

    with pytest.raises(SpecError, match="does not contain sheet"):
        read_workbook(wb_path, workbook_config)


def test_read_workbook_skips_blank_filler_rows(tmp_path: Path) -> None:
    """`Blank` rows are FWF padding bytes - they do not become columns, but
    they extend `record_length` so it matches the actual line width."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    sheet_title = "ImportsSheet"
    ws = wb.create_sheet(sheet_title)
    ws.append(["layout imports", None, None, None, None, None])
    ws.append(["Position", "Description", "Length", "Type", "Nullable", "Parse"])
    ws.append([1, "period", 6, "Date", "n", "yyyymm_to_date"])
    ws.append([7, "Blank", 1, "Char", None, None])  # filler
    ws.append([8, "value", 10, "Num", None, None])
    ws.append([18, "Blank", 3, "Char", None, None])  # trailing filler extends record_length
    wb_path = tmp_path / "wb.xlsx"
    wb.save(wb_path)

    config_path = build_project_config(
        tmp_path / "microtrade.yaml",
        wb_path,
        "2024-01",
        sheet_titles={"imports": sheet_title},
    )
    workbook_config = load_config(config_path).get_workbook(wb_path)

    specs = read_workbook(wb_path, workbook_config)
    imports = specs["imports"]
    assert [c.physical_name for c in imports.columns] == ["period", "value"]
    assert [c.dtype for c in imports.columns] == ["Date", "Int64"]
    assert imports.record_length == 20


def test_read_workbook_honors_position_only_sentinel_row(tmp_path: Path) -> None:
    """A trailing row with Position but no Length marks the end of the record;
    record_length must grow to include that position even though the row
    contributes no real column."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    sheet_title = "ImportsSheet"
    ws = wb.create_sheet(sheet_title)
    ws.append(["Position", "Description", "Length", "Type", "Nullable", "Parse"])
    ws.append([1, "period", 6, "Date", "n", "yyyymm_to_date"])
    ws.append([7, "value", 10, "Num", None, None])
    # Rightmost real column ends at byte 16; sentinel says the record actually
    # extends to byte 17 (an upstream trailing-space filler the workbook
    # declares via a position-only row).
    ws.append([17, None, None, None, None, None])
    wb_path = tmp_path / "wb.xlsx"
    wb.save(wb_path)

    config_path = build_project_config(
        tmp_path / "microtrade.yaml",
        wb_path,
        "2024-01",
        sheet_titles={"imports": sheet_title},
    )
    workbook_config = load_config(config_path).get_workbook(wb_path)

    specs = read_workbook(wb_path, workbook_config)
    imports = specs["imports"]
    assert [c.physical_name for c in imports.columns] == ["period", "value"]
    assert imports.record_length == 17


def test_read_workbook_applies_rename_from_config(schema_workbook: Path, tmp_path: Path) -> None:
    """A `rename` map in microtrade.yaml stamps `logical_name` on matching columns
    so the combined dataset sees a stable logical name even when physical names
    differ across workbook versions."""
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"rename": {"value_usd": "customs_value"}} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    specs = read_workbook(schema_workbook, workbook_config)
    imports = specs["imports"]
    by_physical = {c.physical_name: c for c in imports.columns}
    assert by_physical["value_usd"].logical_name == "customs_value"
    assert by_physical["value_usd"].effective_name == "customs_value"
    # Other imports columns untouched.
    assert by_physical["period"].logical_name is None
    assert by_physical["period"].effective_name == "period"
    # Sheets without a rename map remain logical_name-free.
    for col in specs["exports_us"].columns:
        assert col.logical_name is None


def test_read_workbook_applies_cast_from_config(schema_workbook: Path, tmp_path: Path) -> None:
    """`cast` overrides the workbook's declared dtype per column, and re-derives
    the default parse so e.g. a column cast to Date automatically gets
    `yyyymmdd_to_date`."""
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **(
                            {"cast": {"country_coo": "Utf8", "value_usd": "Float64"}}
                            if tt == "imports"
                            else {}
                        ),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    imports = read_workbook(schema_workbook, workbook_config)["imports"]
    by_physical = {c.physical_name: c for c in imports.columns}
    assert by_physical["value_usd"].dtype == "Float64"  # was Int64, now Float64
    assert by_physical["country_coo"].dtype == "Utf8"  # no-op (already Utf8)
    # Other imports columns untouched.
    assert by_physical["qty_kg"].dtype == "Int64"


def test_read_workbook_parse_override_targets_date_columns(
    schema_workbook: Path, tmp_path: Path
) -> None:
    """`parse` overrides the default parser for a Date column. Casting `period`
    to Date picks up `yyyymmdd_to_date` by default; the `parse` block swaps
    in `yyyymm_to_date` for that single column."""
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **(
                            {
                                "cast": {"period": "Date"},
                                "parse": {"period": "yyyymm_to_date"},
                            }
                            if tt == "imports"
                            else {}
                        ),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    imports = read_workbook(schema_workbook, workbook_config)["imports"]
    period = next(c for c in imports.columns if c.physical_name == "period")
    assert period.dtype == "Date"
    assert period.parse == "yyyymm_to_date"


def test_parse_override_on_non_date_column_raises(schema_workbook: Path, tmp_path: Path) -> None:
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"parse": {"hs_code": "yyyymmdd_to_date"}} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    with pytest.raises(SpecError, match="only meaningful for Date columns"):
        read_workbook(schema_workbook, workbook_config)


def test_cast_rejects_unknown_dtype(schema_workbook: Path, tmp_path: Path) -> None:
    import yaml as yaml_

    from microtrade.config import ConfigError

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"cast": {"value_usd": "Decimal"}} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")

    with pytest.raises(ConfigError, match="are not canonical dtypes"):
        load_config(config_path)


def test_read_workbook_rejects_cast_for_unknown_column(
    schema_workbook: Path, tmp_path: Path
) -> None:
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"cast": {"no_such_col": "Int64"}} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    with pytest.raises(SpecError, match="cast refers to unknown physical column"):
        read_workbook(schema_workbook, workbook_config)


def test_computed_concat_to_date_end_to_end(tmp_path: Path) -> None:
    """A `computed: {entry_date: {kind: concat_to_date, sources: [period, day]}}`
    in the config materializes a Date column at ingest from a YYYYMM period
    and a DD integer."""
    import zipfile
    from datetime import date

    import polars as pl
    import yaml as yaml_
    from openpyxl import Workbook

    from microtrade import pipeline
    from microtrade.discover import RawInput
    from microtrade.ingest import iter_record_batches

    # Workbook: period (6 chars, YYYYMM) + day (2 chars) + value (3 chars).
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("ImportsSheet")
    ws.append(["Position", "Description", "Length", "Type"])
    ws.append([1, "period", 6, "Char"])
    ws.append([7, "day", 2, "Char"])
    ws.append([9, "value", 3, "Char"])
    workbook_path = tmp_path / "wb.xlsx"
    wb.save(workbook_path)

    cfg = {
        "workbooks": {
            workbook_path.name: {
                "effective_from": "2020-01",
                "sheets": {
                    "ImportsSheet": {
                        "trade_type": "imports",
                        "filename_pattern": default_filename_pattern("ImportsSheet"),
                        "cast": {"period": "Date", "day": "Int64"},
                        "parse": {"period": "yyyymm_to_date"},
                        "computed": {
                            "entry_date": {
                                "kind": "concat_to_date",
                                "sources": ["period", "day"],
                            }
                        },
                    }
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(workbook_path)
    imports = read_workbook(workbook_path, workbook_config)["imports"]

    # Spec carries computed_columns and validate_spec passed.
    assert len(imports.computed_columns) == 1
    comp = imports.computed_columns[0]
    assert comp.name == "entry_date" and comp.dtype == "Date"
    assert comp.sources == ("period", "day")

    # Synthesize FWF data and ingest end-to-end.
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    zip_path = input_dir / "ImportsSheet_202406N.TXT.zip"
    # period=YYYYMM(6) day=DD(2) value=(3) -> record_length 11
    lines = ["20240115AAA", "20240315BBB"]
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("data.fwf", "\n".join(lines) + "\n")

    raw = RawInput("imports", 2024, 6, zip_path)
    (batch,) = list(iter_record_batches(raw, imports, chunk_rows=100, encoding="utf-8"))
    assert batch.column("entry_date").to_pylist() == [date(2024, 1, 15), date(2024, 3, 15)]
    assert batch.column("period").to_pylist() == [date(2024, 1, 1), date(2024, 3, 1)]

    # Full pipeline run produces a parquet with the computed column.
    spec_dir = tmp_path / "specs"
    schema.save_spec(imports, spec_dir / "imports" / "v2020-01.yaml")
    output_dir = tmp_path / "output"
    summary = pipeline.run(
        pipeline.PipelineConfig(
            input_dir=input_dir,
            output_dir=output_dir,
            spec_dir=spec_dir,
            trade_types=("imports",),
            ytd=False,
            year=2024,
            month=6,
            encoding="utf-8",
        )
    )
    assert summary.failed_count == 0
    df = pl.scan_parquet(output_dir / "imports" / "**/*.parquet", hive_partitioning=True).collect()
    assert "entry_date" in df.columns
    assert df["entry_date"].to_list() == [date(2024, 1, 15), date(2024, 3, 15)]


def test_computed_invalid_day_goes_to_quality_log(tmp_path: Path) -> None:
    """A day that can't form a valid date (e.g. Feb 30) is routed to the quality
    log as a row-level failure; the surrounding partition still writes."""
    import zipfile

    import yaml as yaml_
    from openpyxl import Workbook

    from microtrade.discover import RawInput
    from microtrade.ingest import QualityIssue, iter_record_batches

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("ImportsSheet")
    ws.append(["Position", "Description", "Length", "Type"])
    ws.append([1, "period", 6, "Char"])
    ws.append([7, "day", 2, "Char"])
    workbook_path = tmp_path / "wb.xlsx"
    wb.save(workbook_path)

    cfg = {
        "workbooks": {
            workbook_path.name: {
                "effective_from": "2020-01",
                "sheets": {
                    "ImportsSheet": {
                        "trade_type": "imports",
                        "filename_pattern": default_filename_pattern("ImportsSheet"),
                        "cast": {"period": "Date", "day": "Int64"},
                        "parse": {"period": "yyyymm_to_date"},
                        "computed": {
                            "entry_date": {
                                "kind": "concat_to_date",
                                "sources": ["period", "day"],
                            }
                        },
                    }
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(workbook_path)
    imports = read_workbook(workbook_path, workbook_config)["imports"]

    zip_path = tmp_path / "ImportsSheet_202406N.TXT.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("data.fwf", "20240215\n20240230\n20240310\n")  # second row is invalid

    raw = RawInput("imports", 2024, 6, zip_path)
    captured: list[QualityIssue] = []
    batches = list(
        iter_record_batches(
            raw, imports, chunk_rows=100, encoding="utf-8", on_quality_issue=captured.append
        )
    )
    # 2 rows land (Feb 15 + Mar 10); the Feb 30 row goes to the quality log.
    total_rows = sum(b.num_rows for b in batches)
    assert total_rows == 2
    assert len(captured) == 1
    assert captured[0].column == "entry_date"
    assert "day is out of range" in captured[0].error or "Feb" in captured[0].error


def test_drop_removes_column_after_computed_uses_it(tmp_path: Path) -> None:
    """A source column dropped via `drop:` is still parsed so a computed column
    can reference it, but doesn't appear in the parquet output."""
    import zipfile
    from datetime import date

    import polars as pl
    import yaml as yaml_
    from openpyxl import Workbook

    from microtrade import pipeline

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("ImportsSheet")
    ws.append(["Position", "Description", "Length", "Type"])
    ws.append([1, "period", 6, "Char"])
    ws.append([7, "day", 2, "Char"])
    workbook_path = tmp_path / "wb.xlsx"
    wb.save(workbook_path)

    cfg = {
        "workbooks": {
            workbook_path.name: {
                "effective_from": "2020-01",
                "sheets": {
                    "ImportsSheet": {
                        "trade_type": "imports",
                        "filename_pattern": default_filename_pattern("ImportsSheet"),
                        "cast": {"period": "Date", "day": "Int64"},
                        "parse": {"period": "yyyymm_to_date"},
                        "computed": {
                            "entry_date": {
                                "kind": "concat_to_date",
                                "sources": ["period", "day"],
                            }
                        },
                        "drop": ["day"],
                    }
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(workbook_path)
    imports = read_workbook(workbook_path, workbook_config)["imports"]
    assert imports.dropped_columns == ("day",)

    # End-to-end: the dropped columns don't appear in the output parquet, but
    # entry_date (which used them) does.
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    zip_path = input_dir / "ImportsSheet_202406N.TXT.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("data.fwf", "20240115\n20240315\n")

    spec_dir = tmp_path / "specs"
    schema.save_spec(imports, spec_dir / "imports" / "v2020-01.yaml")
    output_dir = tmp_path / "output"
    summary = pipeline.run(
        pipeline.PipelineConfig(
            input_dir=input_dir,
            output_dir=output_dir,
            spec_dir=spec_dir,
            trade_types=("imports",),
            ytd=False,
            year=2024,
            month=6,
            encoding="utf-8",
        )
    )
    assert summary.failed_count == 0
    df = pl.scan_parquet(output_dir / "imports" / "**/*.parquet", hive_partitioning=True).collect()
    assert "entry_date" in df.columns
    assert "day" not in df.columns
    # period stays — it's the routing key. `drop` refuses to remove it.
    assert "period" in df.columns
    assert sorted(df["entry_date"].to_list()) == [date(2024, 1, 15), date(2024, 3, 15)]


def test_drop_rejects_unknown_column_name(schema_workbook: Path, tmp_path: Path) -> None:
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"drop": ["no_such_col"]} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    with pytest.raises(SpecError, match=r"dropped_columns .* are not columns"):
        read_workbook(schema_workbook, workbook_config)


def test_drop_rejects_emptying_the_schema(schema_workbook: Path, tmp_path: Path) -> None:
    import yaml as yaml_

    all_names = ["period", "hs_code", "country_coo", "district_entry", "value_usd", "qty_kg"]
    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"drop": all_names} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    with pytest.raises(SpecError, match="leave the output schema empty"):
        read_workbook(schema_workbook, workbook_config)


def test_read_workbook_rejects_rename_for_unknown_column(
    schema_workbook: Path, tmp_path: Path
) -> None:
    import yaml as yaml_

    cfg = {
        "workbooks": {
            schema_workbook.name: {
                "effective_from": "2020-01",
                "sheets": {
                    sheet_title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(sheet_title),
                        **({"rename": {"no_such_column": "something"}} if tt == "imports" else {}),
                    }
                    for tt, sheet_title in SHEET_TITLES.items()
                },
            }
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    workbook_config = load_config(config_path).get_workbook(schema_workbook)

    with pytest.raises(SpecError, match="unknown physical column"):
        read_workbook(schema_workbook, workbook_config)


def test_canonical_columns_merges_on_logical_name(
    schema_workbook: Path, workbook_config: WorkbookConfig
) -> None:
    """Across spec versions, declaring `logical_name` on the newer spec
    pointing at the older spec's physical_name collapses the two into one
    canonical column. Models the real workflow: upstream renames `value_usd`
    -> `customs_value`; the user sets `logical_name=value_usd` on the new
    spec so the combined dataset keeps the stable name."""
    from dataclasses import replace

    imports_v1 = read_workbook(schema_workbook, workbook_config)["imports"]
    imports_v1 = replace(imports_v1, effective_to="2022-12")

    # Simulate v2: upstream renamed `value_usd` to `customs_value`, but the
    # logical_name carries the old name forward.
    new_columns = tuple(
        replace(col, physical_name="customs_value", logical_name="value_usd")
        if col.physical_name == "value_usd"
        else col
        for col in imports_v1.columns
    )
    imports_v2 = replace(
        imports_v1,
        effective_from="2023-01",
        effective_to=None,
        version="2023-01",
        columns=new_columns,
    )

    canonical = schema.canonical_columns([imports_v1, imports_v2])
    names = [c.name for c in canonical]
    assert "value_usd" in names
    assert "customs_value" not in names
    assert len(names) == len(imports_v1.columns)


def test_read_workbook_bakes_effective_to(tmp_path: Path) -> None:
    workbook = build_workbook(tmp_path / "wb.xlsx")
    config_path = build_project_config(
        tmp_path / "microtrade.yaml", workbook, "2020-01", effective_to="2023-12"
    )
    workbook_config = load_config(config_path).get_workbook(workbook)

    specs = read_workbook(workbook, workbook_config)
    for spec in specs.values():
        assert spec.effective_to == "2023-12"


def test_read_workbook_default_workbook_id_derives_from_filename(
    schema_workbook: Path, workbook_config: WorkbookConfig
) -> None:
    specs = read_workbook(schema_workbook, workbook_config)
    for spec in specs.values():
        assert spec.source is not None
        # `schema_workbook.xlsx` -> "schema" (config does not set workbook_id).
        assert spec.source.workbook_id == "schema"


def test_read_workbook_config_workbook_id_wins(tmp_path: Path) -> None:
    workbook = build_workbook(tmp_path / "wb.xlsx")
    config_path = build_project_config(
        tmp_path / "microtrade.yaml", workbook, "2020-01", workbook_id="XYZ12345"
    )
    workbook_config = load_config(config_path).get_workbook(workbook)

    specs = read_workbook(workbook, workbook_config)
    for spec in specs.values():
        assert spec.source is not None
        assert spec.source.workbook_id == "XYZ12345"


# --- import-spec CLI --------------------------------------------------------


def _invoke_import(
    runner: CliRunner, workbook: Path, config_path: Path, out_dir: Path, *extra: str
):
    return runner.invoke(
        app,
        [
            "import-spec",
            str(workbook),
            "--config",
            str(config_path),
            "--out",
            str(out_dir),
            *extra,
        ],
    )


def test_import_spec_cli_writes_yaml(
    schema_workbook: Path, project_config_path: Path, tmp_path: Path
) -> None:
    out_dir = tmp_path / "specs"
    result = _invoke_import(CliRunner(), schema_workbook, project_config_path, out_dir)
    assert result.exit_code == 0, result.output

    for trade_type in TRADE_TYPES:
        path = out_dir / trade_type / "v2020-01.yaml"
        assert path.exists()
        spec = load_spec(path)
        assert spec.trade_type == trade_type
        assert spec.effective_from == "2020-01"
        assert spec.source is not None
        assert spec.source.filename_pattern is not None


def test_import_spec_cli_prints_diff_against_previous(
    schema_workbook: Path, tmp_path: Path
) -> None:
    out_dir = tmp_path / "specs"
    runner = CliRunner()

    cfg_v1 = build_project_config(tmp_path / "v1.yaml", schema_workbook, "2020-01")
    first = _invoke_import(runner, schema_workbook, cfg_v1, out_dir)
    assert first.exit_code == 0

    cfg_v2 = build_project_config(tmp_path / "v2.yaml", schema_workbook, "2025-01")
    second = _invoke_import(runner, schema_workbook, cfg_v2, out_dir)
    assert second.exit_code == 0, second.output
    assert "diff vs v2020-01" in second.output


def test_import_spec_cli_refuses_overwrite_without_force(
    schema_workbook: Path, project_config_path: Path, tmp_path: Path
) -> None:
    out_dir = tmp_path / "specs"
    runner = CliRunner()

    first = _invoke_import(runner, schema_workbook, project_config_path, out_dir)
    assert first.exit_code == 0

    result = _invoke_import(runner, schema_workbook, project_config_path, out_dir)
    assert result.exit_code == 1
    assert "already exists" in result.output

    forced = _invoke_import(runner, schema_workbook, project_config_path, out_dir, "--force")
    assert forced.exit_code == 0


def test_import_spec_cli_errors_when_workbook_missing_from_config(
    schema_workbook: Path, tmp_path: Path
) -> None:
    """If the config lists no entry for this workbook, fail with a clear message."""
    other_workbook = build_workbook(tmp_path / "other.xlsx")
    config_path = build_project_config(tmp_path / "microtrade.yaml", other_workbook, "2020-01")
    out_dir = tmp_path / "specs"

    result = _invoke_import(CliRunner(), schema_workbook, config_path, out_dir)
    # Per-workbook failure -> exit 1 (aggregated), not 2 (config-level error).
    assert result.exit_code == 1
    assert "not listed in the project config" in result.output
    assert "1 of 1 workbook(s) failed" in result.output


def test_import_spec_cli_accepts_multiple_workbooks(tmp_path: Path) -> None:
    """Variadic: glob-style multi-workbook import. Each workbook must have a
    matching entry in the shared microtrade.yaml."""
    import yaml as yaml_

    wb_a = build_workbook(tmp_path / "first.xlsx")
    wb_b = build_workbook(tmp_path / "second.xlsx")
    out_dir = tmp_path / "specs"

    # Shared config: distinct effective_from per workbook so both land.
    cfg = {
        "workbooks": {
            "first.xlsx": {
                "effective_from": "2020-01",
                "effective_to": "2022-12",
                "sheets": {
                    title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(title),
                    }
                    for tt, title in SHEET_TITLES.items()
                },
            },
            "second.xlsx": {
                "effective_from": "2023-01",
                "sheets": {
                    title: {
                        "trade_type": tt,
                        "filename_pattern": default_filename_pattern(title),
                    }
                    for tt, title in SHEET_TITLES.items()
                },
            },
        }
    }
    config_path = tmp_path / "microtrade.yaml"
    config_path.write_text(yaml_.safe_dump(cfg, sort_keys=False), encoding="utf-8")

    result = CliRunner().invoke(
        app,
        [
            "import-spec",
            str(wb_a),
            str(wb_b),
            "--config",
            str(config_path),
            "--out",
            str(out_dir),
        ],
    )
    assert result.exit_code == 0, result.output

    # Both effective_from folders exist per trade type.
    for trade_type in TRADE_TYPES:
        for effective in ("2020-01", "2023-01"):
            assert (out_dir / trade_type / f"v{effective}.yaml").exists()


def test_import_spec_cli_aggregates_failures(schema_workbook: Path, tmp_path: Path) -> None:
    """A bad workbook in the batch doesn't halt the others; command exits 1 at the end."""
    bogus = build_workbook(tmp_path / "bogus.xlsx")
    config_path = build_project_config(tmp_path / "microtrade.yaml", schema_workbook, "2020-01")
    out_dir = tmp_path / "specs"

    result = CliRunner().invoke(
        app,
        [
            "import-spec",
            str(schema_workbook),  # listed in config -> succeeds
            str(bogus),  # not listed -> fails
            "--config",
            str(config_path),
            "--out",
            str(out_dir),
        ],
    )
    assert result.exit_code == 1
    assert (out_dir / "imports" / "v2020-01.yaml").exists()  # first workbook wrote specs
    assert "1 of 2 workbook(s) failed" in result.output


# --- validate-specs ---------------------------------------------------------


def _seed_valid_specs(
    spec_dir: Path, schema_workbook: Path, workbook_config: WorkbookConfig
) -> None:
    specs = read_workbook(schema_workbook, workbook_config)
    for trade_type, spec in specs.items():
        schema.save_spec(spec, spec_dir / trade_type / f"v{spec.effective_from}.yaml")


def test_validate_specs_ok_on_clean_tree(schema_workbook: Path, tmp_path: Path) -> None:
    spec_dir = tmp_path / "specs"
    cfg_v1 = build_project_config(
        tmp_path / "v1.yaml", schema_workbook, "2020-01", effective_to="2023-12"
    )
    cfg_v2 = build_project_config(tmp_path / "v2.yaml", schema_workbook, "2024-01")
    _seed_valid_specs(spec_dir, schema_workbook, load_config(cfg_v1).get_workbook(schema_workbook))
    _seed_valid_specs(spec_dir, schema_workbook, load_config(cfg_v2).get_workbook(schema_workbook))

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 0, result.output
    assert "OK (3 trade types, 6 specs)" in result.output
    assert "imports:" in result.output
    assert "v2020-01" in result.output
    assert "v2024-01" in result.output
    assert "no column changes" in result.output


def test_validate_specs_reports_invalid_yaml(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    _seed_valid_specs(spec_dir, schema_workbook, workbook_config)
    bad = spec_dir / "imports" / "v2020-01.yaml"
    bad.write_text(
        "trade_type: imports\n"
        "version: '2020-01'\n"
        "effective_from: '2020-01'\n"
        "record_length: 10\n"
        "columns:\n"
        "  - {physical_name: a, start: 1, length: 5, dtype: Utf8}\n"
        "  - {physical_name: b, start: 4, length: 5, dtype: Utf8}\n",
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "FAIL" in result.output
    assert "overlaps" in result.output
    assert str(bad) in result.output


def test_validate_specs_rejects_filename_version_mismatch(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    _seed_valid_specs(spec_dir, schema_workbook, workbook_config)
    (spec_dir / "imports" / "v2020-01.yaml").rename(spec_dir / "imports" / "v2020-02.yaml")

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "does not match effective_from" in result.output


def test_validate_specs_flags_window_overlap(schema_workbook: Path, tmp_path: Path) -> None:
    """Two specs whose [effective_from, effective_to] windows overlap must fail."""
    spec_dir = tmp_path / "specs"
    cfg_early = build_project_config(
        tmp_path / "early.yaml", schema_workbook, "2020-01", effective_to="2024-12"
    )
    cfg_late = build_project_config(tmp_path / "late.yaml", schema_workbook, "2024-06")
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_early).get_workbook(schema_workbook)
    )
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_late).get_workbook(schema_workbook)
    )

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "overlapping" in result.output.lower()


def test_validate_specs_flags_gap_between_windows(schema_workbook: Path, tmp_path: Path) -> None:
    spec_dir = tmp_path / "specs"
    cfg_early = build_project_config(
        tmp_path / "early.yaml", schema_workbook, "2020-01", effective_to="2022-12"
    )
    cfg_late = build_project_config(tmp_path / "late.yaml", schema_workbook, "2024-01")
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_early).get_workbook(schema_workbook)
    )
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_late).get_workbook(schema_workbook)
    )

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "gap" in result.output.lower()


def test_validate_specs_flags_open_ended_before_later_spec(
    schema_workbook: Path, tmp_path: Path
) -> None:
    """An earlier spec without `effective_to` must not coexist with a later spec -
    the active window would be ambiguous."""
    spec_dir = tmp_path / "specs"
    cfg_early = build_project_config(tmp_path / "early.yaml", schema_workbook, "2020-01")
    cfg_late = build_project_config(tmp_path / "late.yaml", schema_workbook, "2024-01")
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_early).get_workbook(schema_workbook)
    )
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_late).get_workbook(schema_workbook)
    )

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "open-ended" in result.output.lower()


def test_validate_specs_reports_dtype_conflict_across_versions(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    # First spec with a closed window.
    cfg_early = build_project_config(
        tmp_path / "early.yaml", schema_workbook, "2020-01", effective_to="2024-05"
    )
    _seed_valid_specs(
        spec_dir, schema_workbook, load_config(cfg_early).get_workbook(schema_workbook)
    )
    # Second spec, later window, but we rewrite its `value_usd` dtype.
    cfg_late = build_project_config(tmp_path / "late.yaml", schema_workbook, "2024-06")
    wbcfg_late = load_config(cfg_late).get_workbook(schema_workbook)
    v2 = read_workbook(schema_workbook, wbcfg_late)["imports"]
    new_cols = tuple(
        schema.Column(
            physical_name=c.physical_name,
            start=c.start,
            length=c.length,
            dtype="Float64" if c.physical_name == "value_usd" else c.dtype,
            nullable=c.nullable,
            parse=c.parse,
            description=c.description,
        )
        for c in v2.columns
    )
    v2_conflicting = schema.Spec(
        trade_type=v2.trade_type,
        version=v2.version,
        effective_from=v2.effective_from,
        effective_to=v2.effective_to,
        record_length=v2.record_length,
        columns=new_cols,
        source=v2.source,
        derived=v2.derived,
        partition_by=v2.partition_by,
    )
    schema.save_spec(v2_conflicting, spec_dir / "imports" / "v2024-06.yaml")
    # Also seed the rest so validate-specs has something to compare.
    for trade_type, spec in read_workbook(schema_workbook, wbcfg_late).items():
        if trade_type == "imports":
            continue
        schema.save_spec(spec, spec_dir / trade_type / "v2024-06.yaml")

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "canonical-schema conflict" in result.output
    assert "value_usd" in result.output


def test_validate_specs_empty_tree_exits_nonzero(tmp_path: Path) -> None:
    spec_dir = tmp_path / "empty-specs"
    spec_dir.mkdir()
    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "no specs found" in result.output


def test_validate_specs_continues_across_trade_types(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    _seed_valid_specs(spec_dir, schema_workbook, workbook_config)
    (spec_dir / "imports" / "v2020-01.yaml").write_text(
        "trade_type: imports\n"
        "version: '2020-01'\n"
        "effective_from: '2020-01'\n"
        "record_length: 10\n"
        "columns:\n"
        "  - {physical_name: a, start: 1, length: 5, dtype: Utf8}\n"
        "  - {physical_name: b, start: 4, length: 5, dtype: Utf8}\n",
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 1
    assert "overlaps" in result.output
    assert "exports_us:" in result.output
    assert "exports_nonus:" in result.output


def test_validate_specs_ignores_non_v_prefixed_yaml(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    _seed_valid_specs(spec_dir, schema_workbook, workbook_config)
    (spec_dir / "imports" / "backup.yaml").write_text("not a spec\n", encoding="utf-8")
    (spec_dir / "imports" / "README.yaml").write_text("also not a spec\n", encoding="utf-8")

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 0, result.output
    assert "OK" in result.output


def test_validate_specs_pluralizes_summary_for_singletons(
    schema_workbook: Path, workbook_config: WorkbookConfig, tmp_path: Path
) -> None:
    spec_dir = tmp_path / "specs"
    imports_spec = read_workbook(schema_workbook, workbook_config)["imports"]
    schema.save_spec(imports_spec, spec_dir / "imports" / "v2020-01.yaml")

    result = CliRunner().invoke(app, ["validate-specs", "--spec-dir", str(spec_dir)])
    assert result.exit_code == 0, result.output
    assert "OK (1 trade type, 1 spec)" in result.output


def test_derive_workbook_id_strips_first_underscore_chunk() -> None:
    assert derive_workbook_id("XYZ12345_Record_Layout.xls") == "XYZ12345"
    assert derive_workbook_id("ABC-1234567_Record_Layout.xls") == "ABC-1234567"
    assert derive_workbook_id("plainname.xlsx") == "plainname"
